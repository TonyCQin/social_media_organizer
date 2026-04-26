import csv
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

src = Path('nlp_processor/examples/hand_labeled_posts.csv')
out = Path('nlp_processor/examples/hand_labeled_posts_clean.csv')
wb = load_workbook(filename=BytesIO(src.read_bytes()), read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
if not rows:
    raise SystemExit('No rows found in hand_labeled_posts.csv')
header = [str(col).strip().lower().replace(' ', '_') if col is not None else '' for col in rows[0]]
header = ['content_type' if h == 'content_ty' else h for h in header]
header.append('is_hand_labeled')
with out.open('w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(header)
    for row in rows[1:]:
        values = list(row)
        if len(values) < len(header)-1:
            values.extend([''] * ((len(header)-1) - len(values)))
        values = values[:len(header)-1]
        values.append('TRUE')
        writer.writerow(values)
print('rows', len(rows)-1)
print('header', header)
print('out', out)
